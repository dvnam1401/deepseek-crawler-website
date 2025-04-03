#!/usr/bin/env python3
import os
import re
import json
import time
import argparse
import asyncio
import csv
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from urllib.parse import urlparse, urljoin
from datetime import datetime
import logging
import sys
import requests
import random
from PIL import Image
import io

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

from playwright.async_api import async_playwright, ElementHandle, Page
from config_playwright import OUTPUT_DIR

# Thiết lập logging với encoding UTF-8
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("product_crawler.log", encoding='utf-8'),
        logging.StreamHandler(sys.stdout)  # Sử dụng stdout thay vì stderr
    ]
)

logger = logging.getLogger("product_crawler")

# Đường dẫn lưu trữ dữ liệu
PRODUCT_OUTPUT_DIR = os.path.join(OUTPUT_DIR, "products")
IMAGES_OUTPUT_DIR = os.path.join(OUTPUT_DIR, "images")

# Cấu hình
BASE_URL = "https://www.bachhoaxanh.com"
MAX_RETRIES = 3
REQUEST_TIMEOUT = 30
SCROLL_PAUSE_TIME = 1  # Thời gian chờ sau mỗi lần cuộn
MIN_DELAY = 1  # Thời gian chờ tối thiểu giữa các request (giây)
MAX_DELAY = 3  # Thời gian chờ tối đa giữa các request (giây)
MAX_IMAGES_PER_PRODUCT = 10  # Số lượng hình ảnh tối đa tải về cho mỗi sản phẩm

# Đảm bảo các thư mục tồn tại
for directory in [OUTPUT_DIR, PRODUCT_OUTPUT_DIR, IMAGES_OUTPUT_DIR]:
    os.makedirs(directory, exist_ok=True)

async def wait_for_page_load(page: Page, timeout: int = 30000):
    """Đợi trang web tải hoàn tất"""
    try:
        # Đợi cho đến khi network không còn hoạt động trong 500ms
        await page.wait_for_load_state("networkidle", timeout=timeout)
        # Đợi cho đến khi DOM đã tải xong
        await page.wait_for_load_state("domcontentloaded", timeout=timeout)
    except Exception as e:
        logger.warning(f"Không thể đợi trang tải hoàn tất: {e}")

async def save_screenshot(page: Page, filename: str):
    """Lưu ảnh chụp màn hình của trang hiện tại"""
    screenshot_dir = os.path.join(OUTPUT_DIR, "screenshots")
    os.makedirs(screenshot_dir, exist_ok=True)
    screenshot_path = os.path.join(screenshot_dir, filename)
    await page.screenshot(path=screenshot_path, full_page=True)
    logger.info(f"Đã lưu ảnh chụp màn hình: {screenshot_path}")
    return screenshot_path

async def extract_product_urls(page: Page, selector: str, limit: int) -> List[str]:
    """Trích xuất danh sách URL sản phẩm từ trang danh sách"""
    urls = []
    try:
        # Thực hiện JavaScript để lấy tất cả href
        urls = await page.evaluate(f"""() => {{
            const productElements = document.querySelectorAll('{selector} a');
            const urls = [];
            for (let i = 0; i < productElements.length; i++) {{
                const href = productElements[i].getAttribute('href');
                if (href && !urls.includes(href)) {{
                    urls.push(href);
                }}
            }}
            return urls.slice(0, {limit});
        }}""")
        
        # Thêm domain vào URL nếu cần
        full_urls = [urljoin(BASE_URL, url) for url in urls]
        return full_urls
    except Exception as e:
        logger.error(f"Lỗi khi trích xuất URL sản phẩm: {e}")
        return []

async def extract_product_info_from_list(product_element: ElementHandle) -> Dict[str, Any]:
    """Trích xuất thông tin cơ bản của sản phẩm từ danh sách sản phẩm"""
    info = {}
    
    try:
        # Lấy tên sản phẩm
        name_element = await product_element.query_selector(".product_name")
        if name_element:
            info["name"] = await name_element.text_content()
            info["name"] = info["name"].strip() if info["name"] else ""
        
        # Lấy giá sau giảm
        price_element = await product_element.query_selector(".product_price")
        if price_element:
            info["discounted_price"] = await price_element.text_content()
            info["discounted_price"] = info["discounted_price"].strip() if info["discounted_price"] else ""
        
        # Lấy giá gốc - Sửa selector không hợp lệ
        original_price_element = await product_element.query_selector(".text-12, .line-through, .price-old")
        if original_price_element:
            info["original_price"] = await original_price_element.text_content()
            info["original_price"] = info["original_price"].strip() if info["original_price"] else ""
        
        # Lấy phần trăm giảm giá - Sửa selector không hợp lệ
        discount_element = await product_element.query_selector(".discount-percent, .percent-discount")
        if discount_element:
            info["discount_percent"] = await discount_element.text_content()
            info["discount_percent"] = info["discount_percent"].strip() if info["discount_percent"] else ""
        
        # Lấy URL sản phẩm
        link_element = await product_element.query_selector("a")
        if link_element:
            href = await link_element.get_attribute("href")
            if href:
                info["product_url"] = urljoin(BASE_URL, href)
        
        # Lấy URL hình ảnh
        img_element = await product_element.query_selector("img")
        if img_element:
            src = await img_element.get_attribute("src")
            if src:
                info["image_url"] = src
    
    except Exception as e:
        logger.error(f"Lỗi khi trích xuất thông tin sản phẩm: {e}")
    
    return info

async def check_for_captcha(page: Page) -> bool:
    """Kiểm tra xem có đang hiển thị captcha hay không"""
    captcha_selectors = [
        ".captcha", 
        "#captcha", 
        "input[name='captcha']",
        "img[alt='captcha']",
        "div:has-text('Vui lòng xác minh bạn không phải là robot')"
    ]
    
    for selector in captcha_selectors:
        try:
            captcha_element = await page.query_selector(selector)
            if captcha_element:
                logger.warning(f"Phát hiện captcha trên trang với selector: {selector}")
                return True
        except Exception as e:
            logger.info(f"Lỗi khi kiểm tra captcha với selector {selector}: {e}")
    
    return False

async def handle_captcha(page: Page) -> bool:
    """Xử lý captcha nếu có"""
    is_captcha = await check_for_captcha(page)
    
    if is_captcha:
        logger.warning("Đang gặp captcha, đợi 30 giây cho người dùng giải...")
        # Lưu ảnh chụp màn hình có captcha
        captcha_timestamp = int(time.time())
        screenshot_path = os.path.join(OUTPUT_DIR, "screenshots", f"captcha_{captcha_timestamp}.png")
        await page.screenshot(path=screenshot_path)
        logger.warning(f"Đã lưu ảnh chụp captcha tại: {screenshot_path}")
        
        # Đợi một khoảng thời gian để người dùng can thiệp
        await asyncio.sleep(30)
        
        # Kiểm tra lại sau khi đợi
        is_still_captcha = await check_for_captcha(page)
        if is_still_captcha:
            logger.error("Vẫn còn captcha sau 30 giây, bỏ qua request này")
            return False
        else:
            logger.info("Captcha đã được giải quyết, tiếp tục crawl")
            return True
    
    return True

async def get_product_details(page: Page, product_url: str) -> Dict[str, Any]:
    """Lấy thông tin chi tiết của sản phẩm từ trang sản phẩm"""
    product_details = {}
    
    try:
        for attempt in range(MAX_RETRIES):
            try:
                await page.goto(product_url, wait_until="domcontentloaded", timeout=30000)
                await wait_for_page_load(page)
                
                # Kiểm tra captcha
                captcha_resolved = await handle_captcha(page)
                if not captcha_resolved:
                    logger.error(f"Không thể xử lý captcha trên trang {product_url}, bỏ qua")
                    return product_details
                
                break  # Nếu thành công, thoát khỏi vòng lặp
            except Exception as e:
                logger.warning(f"Lỗi khi tải trang {product_url} (lần thử {attempt+1}/{MAX_RETRIES}): {e}")
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(2)  # Đợi trước khi thử lại
                else:
                    logger.error(f"Đã thử tải trang {MAX_RETRIES} lần nhưng không thành công")
                    return product_details
        
        # Trích xuất tên sản phẩm từ tiêu đề trang
        try:
            title = await page.title()
            if title:
                # Tách tên sản phẩm từ tiêu đề
                parts = title.split(" | ")
                if len(parts) > 0:
                    product_details["title"] = title
                    product_details["name"] = parts[0].strip()
        except Exception as e:
            logger.warning(f"Lỗi khi lấy tiêu đề trang: {e}")
        
        # Lấy tên sản phẩm từ các phần tử DOM nếu chưa có
        if "name" not in product_details:
            name_selectors = ["h1.product-title", ".product-name", "h1", ".product_name"]
            for selector in name_selectors:
                name_element = await page.query_selector(selector)
                if name_element:
                    product_details["name"] = await name_element.text_content()
                    product_details["name"] = product_details["name"].strip() if product_details["name"] else ""
                    break
        
        # Lấy tên sản phẩm thông qua JavaScript để đảm bảo mã hóa Unicode đúng
        if not product_details.get("name"):
            product_details["name"] = await page.evaluate("""
                () => {
                    const nameElements = [
                        document.querySelector('h1.product-title'), 
                        document.querySelector('.product-name'), 
                        document.querySelector('h1'), 
                        document.querySelector('.product_name')
                    ];
                    for (const el of nameElements) {
                        if (el && el.textContent) {
                            return el.textContent.trim();
                        }
                    }
                    return '';
                }
            """)
        
        # Lấy giá sản phẩm
        price_selectors = [".product-price", ".price", ".product_price"]
        for selector in price_selectors:
            price_element = await page.query_selector(selector)
            if price_element:
                product_details["price"] = await price_element.text_content()
                product_details["price"] = product_details["price"].strip() if product_details["price"] else ""
                break
        
        # Lấy mô tả sản phẩm
        description_selectors = [".product-description", ".description", ".detail-content", ".product-content"]
        for selector in description_selectors:
            description_element = await page.query_selector(selector)
            if description_element:
                product_details["description"] = await description_element.text_content()
                product_details["description"] = product_details["description"].strip() if product_details["description"] else ""
                break
        
        # Lấy thông số kỹ thuật
        specs = {}
        specs_rows = await page.query_selector_all(".specifications tr, .product-specs tr, .product-attributes tr")
        for row in specs_rows:
            cells = await row.query_selector_all("td, th")
            if len(cells) >= 2:
                key = await cells[0].text_content()
                value = await cells[1].text_content()
                if key and key.strip():
                    specs[key.strip()] = value.strip() if value else ""
        
        if specs:
            product_details["specifications"] = specs
        
        # Lấy tất cả hình ảnh sản phẩm
        image_selectors = [
            ".product-image img", 
            ".gallery img", 
            ".product-gallery img", 
            ".product img",
            ".picture img",
            "#product-detail-image img",
            "#product-image img",
            ".gallery-container img",
            ".boxprodetail img",
            ".img-product-big .img-fluid",
            ".xzoom",
            ".product-slider-large img"
        ]
        image_urls = []
        
        for selector in image_selectors:
            image_elements = await page.query_selector_all(selector)
            if image_elements:
                for img in image_elements:
                    src = await img.get_attribute("src")
                    if src and src not in image_urls:  # Tránh trùng lặp URL
                        # Chỉ thêm URL có đuôi hình ảnh phổ biến
                        if any(src.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp']):
                            image_urls.append(src)
                        # Hoặc, nếu không có đuôi file, kiểm tra URL chứa từ khóa hình ảnh
                        elif any(keyword in src.lower() for keyword in ['image', 'photo', 'img', 'upload']):
                            image_urls.append(src)
                
                if image_urls:  # Nếu đã tìm thấy hình ảnh, không cần kiểm tra selector khác
                    break
        
        # Nếu không tìm thấy hình ảnh qua selector, thử dùng JS để tìm tất cả hình ảnh trong trang
        if not image_urls:
            try:
                logger.info("Thử dùng JavaScript để tìm ảnh sản phẩm")
                image_urls = await page.evaluate("""
                    () => {
                        const images = Array.from(document.querySelectorAll('img'));
                        // Tìm các hình ảnh có kích thước lớn (có thể là hình sản phẩm chính)
                        return images
                            .filter(img => {
                                // Lọc ra hình ảnh có kích thước phù hợp và loại bỏ icon, logo
                                const rect = img.getBoundingClientRect();
                                return rect.width > 100 && rect.height > 100 &&
                                       img.src && img.src.length > 0;
                            })
                            .map(img => img.src)
                            .filter((url, index, self) => self.indexOf(url) === index); // Loại bỏ trùng lặp
                    }
                """)
                logger.info(f"Đã tìm thấy {len(image_urls)} hình ảnh qua JavaScript")
            except Exception as e:
                logger.warning(f"Lỗi khi tìm hình ảnh qua JavaScript: {e}")
        
        if image_urls:
            product_details["image_urls"] = image_urls
            logger.info(f"Tìm thấy {len(image_urls)} URL hình ảnh cho sản phẩm")
        else:
            logger.warning("Không tìm thấy URL hình ảnh nào cho sản phẩm")
    
    except Exception as e:
        logger.error(f"Lỗi khi lấy thông tin chi tiết sản phẩm từ {product_url}: {e}")
    
    return product_details

async def download_single_image(image_url: str, file_path: str) -> bool:
    """Tải một hình ảnh từ URL và lưu vào đường dẫn cụ thể"""
    # Kiểm tra nếu file đã tồn tại
    if os.path.exists(file_path):
        return True
    
    # Tạo thư mục cha nếu chưa tồn tại
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    # Tải hình ảnh
    for attempt in range(MAX_RETRIES):
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            }
            response = requests.get(image_url, headers=headers, timeout=REQUEST_TIMEOUT)
            response.raise_for_status()
            
            # Kiểm tra kích thước và loại hình ảnh hợp lệ
            if len(response.content) < 100:  # Quá nhỏ có thể là lỗi
                logger.warning(f"Hình ảnh quá nhỏ ({len(response.content)} bytes): {image_url}")
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(1)
                    continue
                return False
            
            # Lưu file gốc
            with open(file_path, "wb") as f:
                f.write(response.content)
            
            # Tạo thumbnail
            try:
                create_thumbnail(file_path)
            except Exception as e:
                logger.warning(f"Không thể tạo thumbnail cho {file_path}: {e}")
            
            return True
        
        except Exception as e:
            logger.warning(f"Lỗi khi tải hình ảnh {image_url} (lần thử {attempt+1}/{MAX_RETRIES}): {e}")
            await asyncio.sleep(1)
    
    logger.error(f"Không thể tải hình ảnh sau {MAX_RETRIES} lần thử: {image_url}")
    return False

def create_thumbnail(original_path: str, max_size: int = 200) -> str:
    """Tạo thumbnail cho hình ảnh và lưu cùng thư mục với tiền tố 'thumb_'"""
    try:
        # Xác định đường dẫn thumbnail
        path_obj = Path(original_path)
        thumb_path = path_obj.parent / f"thumb_{path_obj.name}"
        
        # Nếu thumbnail đã tồn tại, thoát
        if thumb_path.exists():
            return str(thumb_path)
        
        # Mở và resize hình ảnh
        with Image.open(original_path) as img:
            # Giữ tỷ lệ khung hình
            width, height = img.size
            if width > height:
                new_width = max_size
                new_height = int(height * (max_size / width))
            else:
                new_height = max_size
                new_width = int(width * (max_size / height))
            
            # Resize hình ảnh
            img_resized = img.resize((new_width, new_height), Image.LANCZOS)
            
            # Lưu thumbnail
            img_resized.save(thumb_path)
        
        return str(thumb_path)
    except Exception as e:
        logger.error(f"Lỗi khi tạo thumbnail: {e}")
        return ""

async def create_product_gallery(product: Dict[str, Any], image_files: List[str]) -> str:
    """Tạo trang HTML hiển thị gallery các hình ảnh sản phẩm"""
    if not product or not image_files:
        return ""
    
    product_id = product["id"]
    product_name = product.get("name", "Unknown Product")
    product_url = product.get("product_url", "#")
    subcategory = product.get("subcategory", "other")
    price = product.get("price", "N/A")
    description = product.get("description", "Không có mô tả")
    
    # Đường dẫn thư mục sản phẩm
    product_dir = os.path.join(IMAGES_OUTPUT_DIR, subcategory, product_id)
    if not os.path.exists(product_dir):
        return ""
    
    # Tạo HTML gallery
    html_content = f"""<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{product_name} - Hình ảnh sản phẩm</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; background-color: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; }}
        .product-info {{ background-color: white; padding: 20px; border-radius: 5px; margin-bottom: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .product-info h1 {{ margin-top: 0; color: #333; }}
        .product-info a {{ color: #0066cc; text-decoration: none; }}
        .product-info a:hover {{ text-decoration: underline; }}
        .product-info .description {{ margin-top: 15px; padding: 10px; background: #f9f9f9; border-radius: 5px; }}
        .product-meta {{ display: flex; flex-wrap: wrap; gap: 20px; margin-top: 10px; }}
        .product-meta div {{ flex: 1; min-width: 200px; }}
        .gallery {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(250px, 1fr)); gap: 15px; }}
        .gallery-item {{ background-color: white; border-radius: 5px; padding: 10px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .gallery-item img {{ width: 100%; height: auto; cursor: pointer; border-radius: 3px; }}
        .gallery-item p {{ margin: 10px 0 0; color: #666; font-size: 0.9em; }}
        .modal {{ display: none; position: fixed; z-index: 999; left: 0; top: 0; width: 100%; height: 100%; overflow: auto; background-color: rgba(0,0,0,0.9); }}
        .modal-content {{ margin: auto; display: block; max-width: 90%; max-height: 90vh; }}
        .close {{ position: absolute; top: 15px; right: 35px; color: #f1f1f1; font-size: 40px; font-weight: bold; cursor: pointer; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="product-info">
            <h1>{product_name}</h1>
            <div class="product-meta">
                <div>
                    <p><strong>ID sản phẩm:</strong> {product_id}</p>
                    <p><strong>Danh mục:</strong> {subcategory}</p>
                    <p><strong>Giá:</strong> {price}</p>
                </div>
                <div>
                    <p><strong>URL:</strong> <a href="{product_url}" target="_blank">{product_url}</a></p>
                    <p><strong>Ngày tải:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
            </div>
            <div class="description">
                <h3>Mô tả sản phẩm:</h3>
                <p>{description}</p>
            </div>
        </div>
        
        <div class="gallery">
"""
    
    # Tạo các item trong gallery
    for idx, image_file in enumerate(image_files):
        thumb_path = f"thumb_{image_file}"
        
        html_content += f"""
            <div class="gallery-item">
                <img src="{thumb_path}" alt="Hình {idx+1}" onclick="openModal('{image_file}')">
                <p>Hình {idx+1}: {image_file}</p>
            </div>"""
    
    # Đóng HTML
    html_content += """
        </div>
    </div>
    
    <!-- Modal -->
    <div id="imageModal" class="modal">
        <span class="close" onclick="closeModal()">&times;</span>
        <img class="modal-content" id="modalImage">
    </div>
    
    <script>
        function openModal(imagePath) {
            const modal = document.getElementById('imageModal');
            const modalImg = document.getElementById('modalImage');
            modal.style.display = "flex";
            modalImg.src = imagePath;
        }
        
        function closeModal() {
            document.getElementById('imageModal').style.display = "none";
        }
        
        // Đóng modal khi click vào background
        window.onclick = function(event) {
            const modal = document.getElementById('imageModal');
            if (event.target == modal) {
                modal.style.display = "none";
            }
        }
    </script>
</body>
</html>
"""
    
    # Lưu file HTML
    html_path = os.path.join(product_dir, "index.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    
    return html_path

async def download_product_images(page: Page, product: Dict[str, Any]) -> List[str]:
    """Tải tất cả hình ảnh của một sản phẩm và lưu vào thư mục riêng"""
    if not product.get("image_urls") or not product.get("id"):
        return []
    
    product_id = product["id"]
    product_name = product.get("name", "unknown")
    subcategory = product.get("subcategory", "other")
    image_urls = product["image_urls"]
    
    # Tạo thư mục theo cấu trúc: images/subcategory/product_id
    product_image_dir = os.path.join(IMAGES_OUTPUT_DIR, subcategory, product_id)
    os.makedirs(product_image_dir, exist_ok=True)
    
    # Giới hạn số lượng hình ảnh tải về
    image_urls = image_urls[:MAX_IMAGES_PER_PRODUCT]
    
    # Chuẩn hóa product_name để dùng trong tên file
    safe_product_name = re.sub(r'[\\/*?:"<>|]', "", product_name)
    safe_product_name = safe_product_name.replace(" ", "_")[:50]  # Giới hạn độ dài
    
    # Tạo file README.txt trong thư mục sản phẩm
    readme_path = os.path.join(product_image_dir, "README.txt")
    with open(readme_path, "w", encoding="utf-8") as f:
        f.write(f"Sản phẩm: {product_name}\n")
        f.write(f"ID: {product_id}\n")
        f.write(f"URL: {product.get('product_url', 'N/A')}\n")
        f.write(f"Danh mục: {subcategory}\n")
        f.write(f"Giá: {product.get('price', 'N/A')}\n")
        f.write(f"Ngày tải: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write(f"Danh sách {len(image_urls)} hình ảnh:\n")
        for idx, url in enumerate(image_urls):
            f.write(f"{idx+1}. {url}\n")
    
    # Tải các hình ảnh
    downloaded_images = []
    image_files = []
    for idx, img_url in enumerate(image_urls):
        try:
            # Tạo tên file từ tên sản phẩm, index và URL
            parsed_url = urlparse(img_url)
            file_ext = os.path.splitext(parsed_url.path)[1]
            if not file_ext or len(file_ext) > 5:  # Kiểm tra phần mở rộng hợp lệ
                file_ext = ".jpg"  # Mặc định là JPG
            
            # Tạo tên file có ý nghĩa
            filename = f"{safe_product_name}_{idx+1}{file_ext}"
            file_path = os.path.join(product_image_dir, filename)
            
            # Tải hình ảnh
            success = await download_single_image(img_url, file_path)
            if success:
                # Đường dẫn tương đối để lưu trong JSON
                relative_path = os.path.join(subcategory, product_id, filename)
                downloaded_images.append(relative_path)
                image_files.append(filename)
                logger.info(f"Đã tải hình ảnh {idx+1}/{len(image_urls)} cho sản phẩm {product_id}")
            else:
                logger.warning(f"Không thể tải hình ảnh {idx+1}/{len(image_urls)} cho sản phẩm {product_id}")
            
            # Đợi giữa các lần tải để tránh quá tải server
            await asyncio.sleep(0.5)
            
        except Exception as e:
            logger.error(f"Lỗi khi tải hình ảnh {idx+1} cho sản phẩm {product_id}: {e}")
    
    # Cập nhật file README với kết quả tải
    with open(readme_path, "a", encoding="utf-8") as f:
        f.write(f"\nĐã tải thành công: {len(downloaded_images)}/{len(image_urls)} hình ảnh\n")
    
    # Tạo trang HTML gallery
    if image_files:
        html_path = await create_product_gallery(product, image_files)
        logger.info(f"Đã tạo trang gallery cho sản phẩm {product_id}: {html_path}")
    
    return downloaded_images

async def click_load_more_button(page: Page, max_clicks: int = 3):
    """Click vào nút 'Xem thêm' để load thêm sản phẩm"""
    click_count = 0
    
    for _ in range(max_clicks):
        # Các selector phổ biến cho nút "Xem thêm"
        selectors = [
            ".view-more", 
            ".xem-them", 
            ".xemthem", 
            "button:has-text('Xem thêm')", 
            "a:has-text('Xem thêm')",
            ".btn-xemthem",
            "[class*='view-more']",
            "[class*='xem-them']",
            ".show-more",
            ".load-more"
        ]
        
        is_clicked = False
        
        for selector in selectors:
            try:
                # Kiểm tra xem nút có tồn tại và hiển thị không
                is_visible = await page.is_visible(selector)
                if is_visible:
                    logger.info(f"Tìm thấy nút 'Xem thêm' với selector: {selector}")
                    
                    # Cuộn đến nút để đảm bảo nó nhìn thấy được
                    await page.evaluate(f"""
                        const button = document.querySelector('{selector}');
                        if (button) {{
                            button.scrollIntoView({{ behavior: 'smooth', block: 'center' }});
                        }}
                    """)
                    
                    # Đợi một chút để đảm bảo nút đã hiển thị trong viewport
                    await asyncio.sleep(0.5)
                    
                    # Click vào nút
                    await page.click(selector)
                    click_count += 1
                    logger.info(f"Đã click nút 'Xem thêm' lần {click_count}")
                    
                    # Đợi để trang load thêm nội dung
                    await asyncio.sleep(SCROLL_PAUSE_TIME * 2)  # Đợi lâu hơn sau khi click
                    
                    # Đợi cho đến khi network không còn hoạt động
                    try:
                        await page.wait_for_load_state("networkidle", timeout=5000)
                    except Exception as e:
                        logger.warning(f"Không thể đợi network idle: {e}")
                    
                    # Đánh dấu đã click thành công
                    is_clicked = True
                    break
            except Exception as e:
                logger.info(f"Không thể click nút 'Xem thêm' với selector {selector}: {e}")
        
        # Nếu không tìm thấy nút nào hoặc không thể click, dừng vòng lặp
        if not is_clicked:
            logger.info("Không tìm thấy nút 'Xem thêm' nào có thể click")
            break
    
    return click_count

async def scroll_to_load_more_products(page: Page, times: int = 5, target_products: int = 20):
    """Cuộn trang để load thêm sản phẩm"""
    logger.info(f"Bắt đầu cuộn trang để load thêm sản phẩm (mục tiêu: {target_products} sản phẩm)")
    
    # Lấy chiều cao ban đầu của trang
    total_height = await page.evaluate("document.body.scrollHeight")
    viewport_height = await page.evaluate("window.innerHeight")
    
    # Số lượng sản phẩm hiện tại
    current_products = 0
    
    # Thử nhiều selector để đếm sản phẩm
    product_selectors = [".this-item", ".box_product", ".product-item", ".cate-pro-item", "article.product"]
    
    # Lặp cuộn và kiểm tra
    for i in range(times):
        # Đếm số lượng sản phẩm hiện tại
        for selector in product_selectors:
            count = await page.evaluate(f"document.querySelectorAll('{selector}').length")
            if count > current_products:
                current_products = count
                break
        
        logger.info(f"Đã tìm thấy {current_products} sản phẩm sau {i} lần cuộn")
        
        # Nếu đã đủ sản phẩm thì dừng
        if current_products >= target_products:
            logger.info(f"Đã đạt mục tiêu {target_products} sản phẩm, dừng cuộn")
            break
            
        # Thử click nút "Xem thêm" trước khi cuộn
        click_success = await click_load_more_button(page, max_clicks=1)
        
        if click_success > 0:
            logger.info(f"Đã nhấn nút 'Xem thêm', đợi tải thêm sản phẩm")
            await asyncio.sleep(SCROLL_PAUSE_TIME * 2)  # Đợi lâu hơn sau khi click
        else:
            # Cuộn từng đoạn thay vì cuộn thẳng xuống cuối
            current_position = await page.evaluate("window.pageYOffset")
            scroll_step = viewport_height // 2  # Cuộn nửa màn hình mỗi lần
            
            # Nếu đã ở cuối trang, cuộn thêm chút để kích hoạt lazy loading
            if current_position + viewport_height >= total_height - 10:
                # Cuộn nhẹ lên trên rồi xuống lại để kích hoạt lazy loading
                await page.evaluate(f"window.scrollTo(0, {current_position - 100})")
                await asyncio.sleep(0.5)
                await page.evaluate(f"window.scrollTo(0, {total_height})")
            else:
                # Cuộn xuống từng đoạn
                next_position = current_position + scroll_step
                await page.evaluate(f"window.scrollTo(0, {next_position})")
            
            logger.info(f"Đã cuộn lần {i+1}/{times}")
            
            # Đợi để trang load thêm nội dung
            await asyncio.sleep(SCROLL_PAUSE_TIME)
        
        # Cập nhật lại chiều cao trang
        new_height = await page.evaluate("document.body.scrollHeight")
        if new_height > total_height:
            total_height = new_height
        
        # Đợi cho đến khi network không còn hoạt động
        try:
            await page.wait_for_load_state("networkidle", timeout=3000)
        except Exception as e:
            logger.warning(f"Không thể đợi network idle: {e}")
    
    # Kiểm tra xem có đủ sản phẩm chưa, nếu chưa thì cố gắng click thêm nút "Xem thêm" 
    for selector in product_selectors:
        count = await page.evaluate(f"document.querySelectorAll('{selector}').length")
        if count > current_products:
            current_products = count
    
    if current_products < target_products:
        logger.info(f"Chưa đủ sản phẩm ({current_products}/{target_products}), thử nhấn nút 'Xem thêm' lần cuối")
        await click_load_more_button(page, max_clicks=3)
    
    return current_products

async def crawl_products_from_subcategory(page: Page, subcategory_url: str, products_limit: int = 20) -> List[Dict[str, Any]]:
    """Crawl các sản phẩm từ một subcategory"""
    products = []
    
    try:
        await page.goto(subcategory_url, wait_until="domcontentloaded")
        await wait_for_page_load(page)
        
        # Kiểm tra captcha
        captcha_resolved = await handle_captcha(page)
        if not captcha_resolved:
            logger.error(f"Không thể xử lý captcha trên trang {subcategory_url}, bỏ qua")
            return products
        
        # Lưu ảnh chụp màn hình
        subcategory_name = subcategory_url.split("/")[-1]
        await save_screenshot(page, f"subcategory_{subcategory_name}.png")
        
        # Cuộn trang để load thêm sản phẩm
        max_scroll_attempts = max(5, products_limit // 5)  # Số lần cuộn tối đa dựa trên số lượng sản phẩm cần lấy
        found_products = await scroll_to_load_more_products(page, times=max_scroll_attempts, target_products=products_limit)
        logger.info(f"Sau khi cuộn trang, đã tìm thấy {found_products} sản phẩm")
        
        # Tìm các phần tử sản phẩm
        product_selectors = [".this-item", ".box_product", ".product-item", ".cate-pro-item", "article.product"]
        product_urls = []
        
        for selector in product_selectors:
            urls = await extract_product_urls(page, selector, products_limit)
            if urls and len(urls) > len(product_urls):
                product_urls = urls
                logger.info(f"Tìm thấy {len(urls)} URL sản phẩm với selector: {selector}")
        
        if not product_urls:
            logger.warning(f"Không tìm thấy URL sản phẩm nào trên trang {subcategory_url}")
            await save_screenshot(page, f"no_products_{subcategory_name}.png")
            return products
        
        # Giới hạn số lượng sản phẩm
        product_urls = product_urls[:products_limit]
        logger.info(f"Bắt đầu crawl {len(product_urls)} trang sản phẩm")
        
        for idx, product_url in enumerate(product_urls):
            # Thêm độ trễ ngẫu nhiên
            if idx > 0:
                delay = random.uniform(MIN_DELAY, MAX_DELAY)
                logger.info(f"Đợi {delay:.2f} giây trước khi crawl sản phẩm tiếp theo")
                await asyncio.sleep(delay)
            
            try:
                logger.info(f"Đang crawl sản phẩm {idx+1}/{len(product_urls)}: {product_url}")
                
                # Lấy thông tin chi tiết từ trang sản phẩm
                product_details = await get_product_details(page, product_url)
                
                if product_details:
                    # Thêm URL sản phẩm
                    product_details["product_url"] = product_url
                    
                    # Tạo ID cho sản phẩm
                    product_id = f"product_{len(products) + 1}_{int(time.time())}"
                    product_details["id"] = product_id
                    
                    # Thêm thông tin subcategory
                    parsed_url = urlparse(subcategory_url)
                    subcategory = parsed_url.path.strip("/")
                    product_details["subcategory"] = subcategory
                    
                    # Tải hình ảnh sản phẩm vào thư mục riêng
                    if "image_urls" in product_details and product_details["image_urls"]:
                        logger.info(f"Tải {len(product_details['image_urls'][:MAX_IMAGES_PER_PRODUCT])} hình ảnh cho sản phẩm: {product_details.get('name')}")
                        downloaded_images = await download_product_images(page, product_details)
                        product_details["local_images"] = downloaded_images
                    
                    products.append(product_details)
                    logger.info(f"Đã thu thập thông tin sản phẩm: {product_details.get('name', 'Unknown')}")
                else:
                    logger.warning(f"Không thể lấy thông tin chi tiết cho sản phẩm: {product_url}")
            except Exception as e:
                logger.error(f"Lỗi khi xử lý sản phẩm {product_url}: {e}")
    
    except Exception as e:
        logger.error(f"Lỗi khi crawl sản phẩm từ {subcategory_url}: {e}")
    
    return products

def fix_vietnamese_text(text: str) -> str:
    """Sửa các lỗi phổ biến với tiếng Việt trong văn bản"""
    if not text:
        return text
    
    # Sửa các trường hợp cụ thể
    replacements = [
        ("Ba rá»i", "Ba rọi"),
        ("heo nháº­p kháº©u", "heo nhập khẩu"),
        ("Nga tÃºi", "Nga túi"),
        ("51.570Ä'/300g", "51.570đ/300g"),
        ("BÃ¡ch hoÃ¡ XANH", "Bách hoá XANH"),
        ("giÃ¡ tá»'t", "giá tốt"),
        ("táº¡i", "tại"),
        ("tÃºi", "túi"),
        ("Ä'", "đ"),
        ("áº£", "ả"),
        ("á»¥", "ụ"),
        ("á»™", "ộ"),
        # Thêm nhiều trường hợp cải thiện mã hóa tiếng Việt
        ("Ä", "Đ"),
        ("Æ°", "ư"),
        ("Æ¡", "ơ"),
        ("Ã¡", "á"),
        ("Ã ", "à"),
        ("áº¡", "ạ"),
        ("Ã£", "ã"),
        ("Ã¢", "â"),
        ("áº§", "ầ"),
        ("áº¥", "ấ"),
        ("áº­", "ậ"),
        ("áº©", "ẩ"),
        ("áº«", "ẫ"),
        ("Ã©", "é"),
        ("Ã¨", "è"),
        ("áº¹", "ẹ"),
        ("áº½", "ẻ"),
        ("Ãµ", "õ"),
        ("Ã³", "ó"),
        ("Ã²", "ò"),
        # Thêm cho các ký tự đặc biệt
        ("&quot;", "\""),
        ("&amp;", "&"),
        ("&lt;", "<"),
        ("&gt;", ">"),
        ("BÃO CÃO", "BÁO CÁO"),
        ("TỔNG QUAN", "TỔNG QUAN"),
        ("DỮ LIỆU", "DỮ LIỆU"),
        ("Tổng", "Tổng"),
        ("số", "số"),
        ("sản phẩm", "sản phẩm"),
        ("đã", "đã"),
        ("danh mục", "danh mục"),
        ("hình ảnh", "hình ảnh"),
        ("THỐNG KÊ", "THỐNG KÊ"),
        ("CÁC FILE ĐàTẠO", "CÁC FILE ĐÃ TẠO"),
        ("CÁC FILE ĐÃ TẠO", "CÁC FILE ĐÃ TẠO"),
        ("TẠO", "TẠO")
    ]
    
    result = text
    for old, new in replacements:
        result = result.replace(old, new)
    
    return result

def normalize_product_data(products: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Tiền xử lý dữ liệu sản phẩm để đảm bảo định dạng chuẩn"""
    normalized_products = []
    
    for product in products:
        normalized_product = {}
        
        for key, value in product.items():
            # Chuẩn hóa các chuỗi
            if isinstance(value, str):
                # Sửa lỗi tiếng Việt
                normalized_value = fix_vietnamese_text(value.strip())
                normalized_product[key] = normalized_value
            else:
                normalized_product[key] = value
        
        normalized_products.append(normalized_product)
    
    return normalized_products

def save_products_to_file(products: List[Dict[str, Any]], subcategory_name: str) -> str:
    """Lưu thông tin sản phẩm vào file JSON"""
    if not products:
        logger.warning(f"Không có sản phẩm nào để lưu cho subcategory {subcategory_name}")
        return ""
    
    # Tiền xử lý dữ liệu sản phẩm
    normalized_products = normalize_product_data(products)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{subcategory_name}_{timestamp}.json"
    file_path = os.path.join(PRODUCT_OUTPUT_DIR, filename)
    
    try:
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(normalized_products, f, ensure_ascii=False, indent=2)
        
        logger.info(f"Đã lưu {len(products)} sản phẩm vào file: {file_path}")
        return file_path
    
    except Exception as e:
        logger.error(f"Lỗi khi lưu sản phẩm vào file: {e}")
        return ""

def save_products_to_csv(products: List[Dict[str, Any]], subcategory_name: str) -> str:
    """Lưu thông tin sản phẩm vào file CSV"""
    if not products:
        logger.warning(f"Không có sản phẩm nào để lưu cho subcategory {subcategory_name}")
        return ""
    
    # Tiền xử lý dữ liệu sản phẩm
    normalized_products = normalize_product_data(products)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{subcategory_name}_{timestamp}.csv"
    file_path = os.path.join(PRODUCT_OUTPUT_DIR, filename)
    
    try:
        # Xác định các trường chính để làm header
        all_fields = set()
        for product in normalized_products:
            all_fields.update(product.keys())
        
        # Loại bỏ các trường không muốn xuất ra CSV
        fields_to_exclude = {'image_urls', 'specifications', 'local_images'}
        csv_fields = [field for field in all_fields if field not in fields_to_exclude]
        
        # Thêm các trường quan trọng vào đầu danh sách
        important_fields = ['id', 'name', 'price', 'subcategory', 'product_url']
        for field in reversed(important_fields):
            if field in csv_fields:
                csv_fields.remove(field)
                csv_fields.insert(0, field)
        
        # Sử dụng pandas để xử lý UTF-8 tốt hơn
        df_data = []
        for product in normalized_products:
            row_data = {}
            for field in csv_fields:
                value = product.get(field, "")
                # Đảm bảo mọi giá trị đều là chuỗi và được xử lý tiếng Việt
                if isinstance(value, str):
                    row_data[field] = fix_vietnamese_text(value)
                else:
                    row_data[field] = value
            df_data.append(row_data)
        
        # Dùng pandas để tạo và lưu CSV
        df = pd.DataFrame(df_data)
        df.to_csv(file_path, index=False, encoding='utf-8-sig')  # Dùng UTF-8 với BOM để Excel nhận diện đúng
        
        logger.info(f"Đã lưu {len(products)} sản phẩm vào file CSV: {file_path}")
        return file_path
    
    except Exception as e:
        logger.error(f"Lỗi khi lưu sản phẩm vào file CSV: {e}")
        return ""

def save_products_to_excel(products: List[Dict[str, Any]], subcategory_name: str) -> str:
    """Lưu thông tin sản phẩm vào file Excel với định dạng đẹp"""
    if not EXCEL_SUPPORT:
        logger.warning("Không thể xuất Excel. Hãy cài đặt pandas và openpyxl: pip install pandas openpyxl")
        return ""
    
    if not products:
        logger.warning(f"Không có sản phẩm nào để lưu cho subcategory {subcategory_name}")
        return ""
    
    # Tiền xử lý dữ liệu sản phẩm
    normalized_products = normalize_product_data(products)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{subcategory_name}_{timestamp}.xlsx"
    file_path = os.path.join(PRODUCT_OUTPUT_DIR, filename)
    
    try:
        # Xác định các trường chính để làm header
        all_fields = set()
        for product in normalized_products:
            all_fields.update(product.keys())
        
        # Loại bỏ các trường không muốn xuất ra Excel
        fields_to_exclude = {'image_urls', 'specifications', 'local_images'}
        excel_fields = [field for field in all_fields if field not in fields_to_exclude]
        
        # Thêm các trường quan trọng vào đầu danh sách
        important_fields = ['id', 'name', 'price', 'subcategory', 'product_url', 'description']
        for field in reversed(important_fields):
            if field in excel_fields:
                excel_fields.remove(field)
                excel_fields.insert(0, field)
        
        # Chuyển dữ liệu sang DataFrame
        df_data = []
        for product in normalized_products:
            row_data = {}
            for field in excel_fields:
                value = product.get(field, "")
                # Đảm bảo mọi giá trị đều là chuỗi và được xử lý tiếng Việt
                if isinstance(value, str):
                    row_data[field] = fix_vietnamese_text(value)
                else:
                    row_data[field] = value
            df_data.append(row_data)
        
        df = pd.DataFrame(df_data)
        
        # Tạo writer Excel
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        df.to_excel(writer, sheet_name='Products', index=False)
        
        # Lấy worksheet để định dạng
        workbook = writer.book
        worksheet = writer.sheets['Products']
        
        # Định dạng header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Định dạng border
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Áp dụng định dạng cho header
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Chỉnh sửa chiều rộng cột
        for idx, col in enumerate(df.columns):
            column_width = 20  # Chiều rộng mặc định
            
            # Điều chỉnh chiều rộng dựa trên loại dữ liệu
            if col == 'name':
                column_width = 40
            elif col == 'description':
                column_width = 60
            elif col == 'product_url':
                column_width = 50
            elif col == 'price':
                column_width = 15
            
            column_letter = openpyxl.utils.get_column_letter(idx + 1)
            worksheet.column_dimensions[column_letter].width = column_width
        
        # Định dạng hàng dữ liệu
        data_alignment = Alignment(vertical='top', wrap_text=True)
        for row_num in range(2, len(df) + 2):
            # Chiều cao hàng
            worksheet.row_dimensions[row_num].height = 30
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = data_alignment
                cell.border = thin_border
        
        # Thêm lọc tự động
        worksheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}{len(df) + 1}"
        
        # Đóng băng hàng đầu tiên
        worksheet.freeze_panes = "A2"
        
        # Lưu file Excel
        writer.close()
        
        logger.info(f"Đã lưu {len(products)} sản phẩm vào file Excel: {file_path}")
        return file_path
        
    except Exception as e:
        logger.error(f"Lỗi khi lưu sản phẩm vào file Excel: {e}")
        return ""

def generate_summary_report(all_results: List[Dict[str, Any]], output_dir: str):
    """Tạo báo cáo tổng quan sau khi crawl hoàn tất"""
    if not all_results:
        logger.warning("Không có dữ liệu để tạo báo cáo tổng quan")
        return
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_file = os.path.join(output_dir, f"crawl_summary_{timestamp}.txt")
    
    try:
        # Chuẩn bị nội dung báo cáo trước
        report_content = "="*80 + "\n"
        report_content += f"BÁO CÁO TỔNG QUAN CRAWL DỮ LIỆU - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        report_content += "="*80 + "\n\n"
        
        # Thống kê tổng quát
        total_products = len(all_results)
        subcategories = set(product.get("subcategory", "unknown") for product in all_results)
        total_images = sum(len(product.get("local_images", [])) for product in all_results)
        
        report_content += f"Tổng số sản phẩm đã crawl: {total_products}\n"
        report_content += f"Tổng số danh mục: {len(subcategories)}\n"
        report_content += f"Tổng số hình ảnh đã tải: {total_images}\n\n"
        
        # Thống kê theo danh mục
        report_content += "THỐNG KÊ THEO DANH MỤC:\n"
        report_content += "-"*50 + "\n"
        subcategory_stats = {}
        for product in all_results:
            subcat = product.get("subcategory", "unknown")
            if subcat not in subcategory_stats:
                subcategory_stats[subcat] = {"count": 0, "images": 0}
            subcategory_stats[subcat]["count"] += 1
            subcategory_stats[subcat]["images"] += len(product.get("local_images", []))
        
        for subcat, stats in sorted(subcategory_stats.items(), key=lambda x: x[1]["count"], reverse=True):
            report_content += f"- {subcat}: {stats['count']} sản phẩm, {stats['images']} hình ảnh\n"
        
        report_content += "\n"
        
        # Danh sách các file output đã tạo
        report_content += "CÁC FILE ĐÃ TẠO:\n"
        report_content += "-"*50 + "\n"
        
        # Liệt kê các file JSON
        json_files = [f for f in os.listdir(PRODUCT_OUTPUT_DIR) if f.endswith(".json")]
        if json_files:
            report_content += "Files JSON:\n"
            for json_file in sorted(json_files):
                file_path = os.path.join(PRODUCT_OUTPUT_DIR, json_file)
                file_size = os.path.getsize(file_path) / 1024  # KB
                report_content += f"- {json_file} ({file_size:.2f} KB)\n"
            report_content += "\n"
        
        # Liệt kê các file CSV
        csv_files = [f for f in os.listdir(PRODUCT_OUTPUT_DIR) if f.endswith(".csv")]
        if csv_files:
            report_content += "Files CSV:\n"
            for csv_file in sorted(csv_files):
                file_path = os.path.join(PRODUCT_OUTPUT_DIR, csv_file)
                file_size = os.path.getsize(file_path) / 1024  # KB
                report_content += f"- {csv_file} ({file_size:.2f} KB)\n"
            report_content += "\n"
        
        # Liệt kê các file Excel
        excel_files = [f for f in os.listdir(PRODUCT_OUTPUT_DIR) if f.endswith(".xlsx")]
        if excel_files:
            report_content += "Files Excel:\n"
            for excel_file in sorted(excel_files):
                file_path = os.path.join(PRODUCT_OUTPUT_DIR, excel_file)
                file_size = os.path.getsize(file_path) / 1024  # KB
                report_content += f"- {excel_file} ({file_size:.2f} KB)\n"
            report_content += "\n"
        
        # Thống kê thư mục hình ảnh
        image_dirs = []
        for root, dirs, files in os.walk(IMAGES_OUTPUT_DIR):
            for dir_name in dirs:
                if dir_name.startswith("product_"):
                    image_dirs.append(os.path.join(root, dir_name))
        
        if image_dirs:
            report_content += f"Số lượng thư mục hình ảnh sản phẩm: {len(image_dirs)}\n"
            report_content += "\n"
        
        # Danh sách 10 sản phẩm tiêu biểu (theo số lượng hình ảnh)
        top_products = sorted(all_results, key=lambda x: len(x.get("local_images", [])), reverse=True)[:10]
        if top_products:
            report_content += "TOP 10 SẢN PHẨM CÓ NHIỀU HÌNH ẢNH NHẤT:\n"
            report_content += "-"*50 + "\n"
            for i, product in enumerate(top_products, 1):
                name = fix_vietnamese_text(product.get("name", "Unknown"))
                images_count = len(product.get("local_images", []))
                subcat = product.get("subcategory", "unknown")
                report_content += f"{i}. {name} [{subcat}] - {images_count} hình ảnh\n"
            report_content += "\n"
        
        # Thời gian hoàn thành
        report_content += f"\nBáo cáo được tạo lúc: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        report_content += "="*80 + "\n"
        
        # Sửa tiếng Việt trong toàn bộ báo cáo
        report_content = fix_vietnamese_text(report_content)
        
        # Ghi nội dung đã chuẩn bị vào file
        with open(report_file, "w", encoding="utf-8-sig") as f:
            f.write(report_content)
        
        logger.info(f"Đã tạo báo cáo tổng quan tại: {report_file}")
        
        # In ra terminal phần tóm tắt
        print("\n" + "="*40)
        print("TÓM TẮT KẾT QUẢ CRAWL:")
        print("="*40)
        print(f"- Tổng số sản phẩm đã crawl: {total_products}")
        print(f"- Tổng số danh mục: {len(subcategories)}")
        print(f"- Tổng số hình ảnh đã tải: {total_images}")
        print(f"- Báo cáo chi tiết: {report_file}")
        print("="*40 + "\n")
        
        return report_file
        
    except Exception as e:
        logger.error(f"Lỗi khi tạo báo cáo tổng quan: {e}")
        return None

async def crawl_subcategories(categories_file: str, product_limit: int = 20, subcategory_limit: int = None, export_csv: bool = False, export_excel: bool = False):
    """Quản lý crawl các subcategories"""
    # Đọc danh sách subcategories từ file JSON
    try:
        with open(categories_file, "r", encoding="utf-8") as f:
            categories_data = json.load(f)
        logger.info(f"Đã đọc file categories: {categories_file}")
    except Exception as e:
        logger.error(f"Lỗi khi đọc file categories {categories_file}: {e}")
        return
    
    # Chuẩn bị danh sách các subcategory URLs
    subcategory_urls = []
    for category in categories_data:
        if "subcategories" in category:
            for subcat in category["subcategories"]:
                if "subcategory_url" in subcat:
                    subcategory_urls.append(subcat["subcategory_url"])
    
    # In thông tin debug về subcategories
    logger.info(f"Tìm thấy {len(subcategory_urls)} subcategories từ file")
    if not subcategory_urls:
        logger.warning(f"Cấu trúc file JSON: {categories_data[:2] if categories_data else 'File trống'}")
    
    # Giới hạn số lượng subcategories (nếu có)
    if subcategory_limit and subcategory_limit > 0:
        subcategory_urls = subcategory_urls[:subcategory_limit]
    
    logger.info(f"Chuẩn bị crawl {len(subcategory_urls)} subcategories")
    
    # Nếu không có subcategories, thoát
    if not subcategory_urls:
        logger.error("Không tìm thấy subcategory URLs trong file. Kiểm tra lại file categories.")
        return
    
    # Lưu tất cả kết quả
    all_results = []
    
    # Chạy Playwright
    async with async_playwright() as p:
        # Khởi tạo browser
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            viewport={"width": 1280, "height": 720},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        )
        page = await context.new_page()
        
        total_products = 0
        
        # Duyệt qua từng subcategory
        for subcategory_url in subcategory_urls:
            try:
                subcategory_name = subcategory_url.split("/")[-1]
                logger.info(f"Bắt đầu crawl subcategory: {subcategory_name} - {subcategory_url}")
                
                start_time = time.time()
                
                # Crawl sản phẩm
                products = await crawl_products_from_subcategory(page, subcategory_url, product_limit)
                
                # Lưu sản phẩm vào file và thêm vào danh sách kết quả
                if products:
                    # Thêm vào kết quả tổng hợp
                    all_results.extend(products)
                    
                    # Lưu json mặc định
                    save_products_to_file(products, subcategory_name)
                    
                    # Xuất ra CSV nếu được yêu cầu
                    if export_csv:
                        save_products_to_csv(products, subcategory_name)
                    
                    # Xuất ra Excel nếu được yêu cầu
                    if export_excel:
                        save_products_to_excel(products, subcategory_name)
                    
                    total_products += len(products)
                
                end_time = time.time()
                logger.info(f"Đã crawl {len(products)} sản phẩm từ {subcategory_name} trong {end_time - start_time:.2f} giây")
                
                # Delay để tránh quá tải server
                await asyncio.sleep(2)
                
            except Exception as e:
                logger.error(f"Lỗi khi xử lý subcategory {subcategory_url}: {e}")
        
        # Đóng browser
        await browser.close()
        
        logger.info(f"Hoàn thành quá trình crawl. Tổng cộng: {total_products} sản phẩm từ {len(subcategory_urls)} subcategories")
    
    # Tạo báo cáo tổng quan
    if all_results:
        generate_summary_report(all_results, PRODUCT_OUTPUT_DIR)

async def main():
    parser = argparse.ArgumentParser(description="Crawl thông tin sản phẩm từ bachhoaxanh.com")
    parser.add_argument("--categories", type=str, default="data/categories_playwright.json", help="File JSON chứa danh sách categories")
    parser.add_argument("--products", type=int, default=20, help="Số lượng sản phẩm tối đa crawl từ mỗi subcategory")
    parser.add_argument("--subcategories", type=int, default=None, help="Số lượng subcategories tối đa sẽ crawl")
    parser.add_argument("--csv", action="store_true", help="Xuất dữ liệu dưới dạng CSV")
    parser.add_argument("--excel", action="store_true", help="Xuất dữ liệu dưới dạng Excel")
    
    args = parser.parse_args()
    
    await crawl_subcategories(args.categories, args.products, args.subcategories, args.csv, args.excel)

if __name__ == "__main__":
    asyncio.run(main()) 